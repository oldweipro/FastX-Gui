import json

from loguru import logger

from app.database.repository import UnitOfWork


class ImportExportService:
    """Import/Export projects in JSON and XLSX formats."""

    # ── JSON Export ──────────────────────────────────────────────

    def export_project_json(self, project_id: str, file_path: str) -> None:
        """Export a single project with all templates, fields, items to JSON."""
        data = self._serialize_project(project_id)
        if data is None:
            raise ValueError(f"Project {project_id} not found")
        output = {"version": "2.0", "export_type": "project", **data}
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2, default=str)
        logger.info(f"[ImportExportService] Exported project to {file_path}")

    def export_workspace_json(self, workspace_id: str, file_path: str) -> None:
        """Export an entire workspace with all projects."""
        with UnitOfWork() as uow:
            ws = uow.workspaces.get_by_id(workspace_id)
            if ws is None:
                raise ValueError(f"Workspace {workspace_id} not found")
            projects = uow.projects.get_by_workspace(workspace_id)

        project_data = []
        for p in projects:
            pd = self._serialize_project(p.id)
            if pd:
                project_data.append(pd)

        output = {
            "version": "2.0",
            "export_type": "workspace",
            "workspace": {"id": ws.id, "name": ws.name, "description": ws.description},
            "projects": project_data,
        }
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2, default=str)
        logger.info(f"[ImportExportService] Exported workspace to {file_path}")

    # ── JSON Import ──────────────────────────────────────────────

    def import_project_json(self, workspace_id: str, file_path: str) -> dict:
        """Import project(s) from a JSON file into a workspace. Returns summary."""
        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)

        export_type = data.get("export_type", "project")
        imported = []

        if export_type == "workspace":
            for pd in data.get("projects", []):
                result = self._import_single_project(workspace_id, pd)
                imported.append(result)
        else:
            result = self._import_single_project(workspace_id, data)
            imported.append(result)

        logger.info(f"[ImportExportService] Imported {len(imported)} project(s) from {file_path}")
        return {"imported_projects": imported}

    # ── XLSX Export ───────────────────────────────────────────────

    def export_project_xlsx(self, project_id: str, file_path: str) -> None:
        """Export project to XLSX with multiple sheets."""
        import openpyxl

        data = self._serialize_project(project_id)
        if data is None:
            raise ValueError(f"Project {project_id} not found")

        wb = openpyxl.Workbook()

        # Project sheet
        ws_proj = wb.active
        ws_proj.title = "Project"
        proj = data["project"]
        ws_proj.append(["ID", "Name", "Description"])
        ws_proj.append([proj["id"], proj["name"], proj["description"]])

        # Templates sheet
        ws_tmpl = wb.create_sheet("Templates")
        ws_tmpl.append(["ID", "Name", "Description", "Fields Count"])
        for t in data.get("templates", []):
            ws_tmpl.append([t["id"], t["name"], t.get("description", ""), len(t.get("fields", []))])

        # Fields sheet
        ws_fields = wb.create_sheet("Fields")
        ws_fields.append(["Template", "Field ID", "Name", "Type", "Label", "Required", "Options"])
        for t in data.get("templates", []):
            for f in t.get("fields", []):
                opts = f.get("options", [])
                ws_fields.append([
                    t["name"], f["id"], f["name"], f["field_type"],
                    f.get("label", ""), f.get("required", False),
                    json.dumps(opts, ensure_ascii=False) if opts else "",
                ])

        # Items sheet
        ws_items = wb.create_sheet("Items")
        ws_items.append(["ID", "Template ID", "Title", "Field Values"])
        for item in data.get("items", []):
            fv = item.get("field_values", {})
            ws_items.append([
                item["id"], item["template_id"], item["title"],
                json.dumps(fv, ensure_ascii=False) if fv else "",
            ])

        wb.save(file_path)
        logger.info(f"[ImportExportService] Exported project XLSX to {file_path}")

    # ── XLSX Import ───────────────────────────────────────────────

    def import_project_xlsx(self, workspace_id: str, file_path: str) -> dict:
        """Import project from XLSX file."""
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True)

        # Read project info
        ws_proj = wb["Project"]
        rows = list(ws_proj.iter_rows(min_row=2, values_only=True))
        if not rows:
            raise ValueError("No project data found in XLSX")
        proj_row = rows[0]
        project_name = proj_row[1] if len(proj_row) > 1 else "Imported Project"
        project_desc = proj_row[2] if len(proj_row) > 2 else ""

        # Read templates
        templates = []
        if "Templates" in wb.sheetnames:
            ws_tmpl = wb["Templates"]
            for row in ws_tmpl.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    templates.append({"id": str(row[0]), "name": str(row[1]), "description": str(row[2] or "")})

        # Read fields
        fields_by_tmpl_name: dict[str, list[dict]] = {}
        if "Fields" in wb.sheetnames:
            ws_fields = wb["Fields"]
            for row in ws_fields.iter_rows(min_row=2, values_only=True):
                tmpl_name = str(row[0]) if row[0] else ""
                if tmpl_name not in fields_by_tmpl_name:
                    fields_by_tmpl_name[tmpl_name] = []
                options = []
                if row[6]:
                    try:
                        options = json.loads(str(row[6]))
                    except (json.JSONDecodeError, TypeError):
                        pass
                fields_by_tmpl_name[tmpl_name].append({
                    "name": str(row[2] or ""),
                    "field_type": str(row[3] or "text"),
                    "label": str(row[4] or ""),
                    "required": bool(row[5]) if row[5] else False,
                    "options": options,
                })

        # Read items
        items = []
        if "Items" in wb.sheetnames:
            ws_items = wb["Items"]
            for row in ws_items.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    fv = {}
                    if row[3]:
                        try:
                            fv = json.loads(str(row[3]))
                        except (json.JSONDecodeError, TypeError):
                            pass
                    items.append({
                        "old_template_id": str(row[1]) if row[1] else "",
                        "title": str(row[2] or ""),
                        "field_values": fv,
                    })

        wb.close()

        # Create project and data with new IDs
        id_map = {}  # old_template_id -> new_template_id
        with UnitOfWork() as uow:
            proj = uow.projects.create(workspace_id=workspace_id, name=project_name, description=project_desc)

            for td in templates:
                tmpl = uow.templates.create(project_id=proj.id, name=td["name"], description=td.get("description", ""))
                id_map[td["id"]] = tmpl.id
                for fd in fields_by_tmpl_name.get(td["name"], []):
                    uow.fields.create(template_id=tmpl.id, **fd)

            for id_item in items:
                old_tid = id_item.get("old_template_id", "")
                new_tid = id_map.get(old_tid, old_tid)
                uow.items.create(
                    project_id=proj.id,
                    template_id=new_tid,
                    title=id_item["title"],
                    field_values=id_item.get("field_values", {}),
                )
            uow.commit()

        logger.info(f"[ImportExportService] Imported XLSX project: {project_name}")
        return {"imported_projects": [{"name": project_name, "templates": len(templates), "items": len(items)}]}

    # ── Internal helpers ─────────────────────────────────────────

    def _serialize_project(self, project_id: str) -> dict | None:
        with UnitOfWork() as uow:
            proj = uow.projects.get_by_id(project_id)
            if proj is None:
                return None

            templates_data = []
            for t in uow.templates.get_by_project(project_id):
                fields = []
                for f in uow.fields.get_by_template(t.id):
                    options = f.options
                    if isinstance(options, str):
                        try:
                            options = json.loads(options)
                        except (json.JSONDecodeError, TypeError):
                            options = []
                    fields.append({
                        "id": f.id,
                        "name": f.name,
                        "field_type": f.field_type,
                        "label": f.label,
                        "required": f.required,
                        "default_val": f.default_val,
                        "options": options,
                        "ref_tmpl_id": f.ref_tmpl_id,
                        "multi_select": f.multi_select,
                        "sort_order": f.sort_order,
                    })
                templates_data.append({
                    "id": t.id,
                    "name": t.name,
                    "description": t.description,
                    "sort_order": t.sort_order,
                    "fields": fields,
                    "sub_templates": [st.id for st in t.sub_templates],
                })

            items_data = []
            for item in uow.items.get_by_project(project_id):
                fv = item.field_values
                if isinstance(fv, str):
                    try:
                        fv = json.loads(fv)
                    except (json.JSONDecodeError, TypeError):
                        fv = {}
                items_data.append({
                    "id": item.id,
                    "template_id": item.template_id,
                    "title": item.title,
                    "field_values": fv,
                })

            return {
                "project": {"id": proj.id, "name": proj.name, "description": proj.description},
                "templates": templates_data,
                "items": items_data,
            }

    def _import_single_project(self, workspace_id: str, data: dict) -> dict:
        """Import one project from parsed JSON data with ID remapping."""
        proj_info = data.get("project", {})
        proj_name = proj_info.get("name", "Imported Project")
        proj_desc = proj_info.get("description", "")

        id_map = {}  # old_id -> new_id

        with UnitOfWork() as uow:
            proj = uow.projects.create(workspace_id=workspace_id, name=proj_name, description=proj_desc)
            id_map[proj_info.get("id", "")] = proj.id

            # Create templates and fields
            for td in data.get("templates", []):
                tmpl = uow.templates.create(project_id=proj.id, name=td["name"], description=td.get("description", ""))
                old_tmpl_id = td.get("id", "")
                id_map[old_tmpl_id] = tmpl.id
                tmpl.sort_order = td.get("sort_order", 0)

                for fd in td.get("fields", []):
                    old_field_id = fd.get("id", "")
                    ref = fd.get("ref_tmpl_id")
                    field = uow.fields.create(
                        template_id=tmpl.id,
                        name=fd["name"],
                        field_type=fd["field_type"],
                        label=fd.get("label", ""),
                        required=fd.get("required", False),
                        default_val=fd.get("default_val", ""),
                        options=fd.get("options", []),
                        ref_tmpl_id=id_map.get(ref, ref) if ref else None,
                        multi_select=fd.get("multi_select", False),
                        sort_order=fd.get("sort_order", 0),
                    )
                    id_map[old_field_id] = field.id

            # Restore sub-template relationships with remapped IDs
            for td in data.get("templates", []):
                new_parent_id = id_map.get(td.get("id", ""))
                if new_parent_id:
                    for sub_id in td.get("sub_templates", []):
                        new_sub_id = id_map.get(sub_id)
                        if new_sub_id:
                            try:
                                uow.templates.add_sub_template(new_parent_id, new_sub_id)
                            except Exception:
                                pass

            # Create items with remapped template IDs
            for id_data in data.get("items", []):
                old_tid = id_data.get("template_id", "")
                new_tid = id_map.get(old_tid, old_tid)
                uow.items.create(
                    project_id=proj.id,
                    template_id=new_tid,
                    title=id_data.get("title", ""),
                    field_values=id_data.get("field_values", {}),
                )

            uow.commit()

        return {
            "name": proj_name,
            "templates": len(data.get("templates", [])),
            "items": len(data.get("items", [])),
        }
